from openpyxl import load_workbook
import pandas as pd
import logging

class ExcelReader:
    def __init__ (self, file_path : str):
        self.logger = logging.getLogger('PortfolioAnalyzer')        
        self.file_path = file_path
        self.workbook = load_workbook(filename=file_path, data_only=True)

    def __delattr__(self, name: str) -> None:
        self.workbook.close()        

    def read_table(self, sheet_name : str, table_name : str) -> pd.DataFrame:
        if sheet_name not in self.workbook.sheetnames:    
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")
        
        sheet = self.workbook[sheet_name]
        
        table_range = None
        if hasattr(sheet, 'tables') and table_name in sheet.tables:
            table = sheet.tables[table_name]
            table_range = table.ref
            self.logger.debug(f"Found Excel table '{table_name}' at range: {table_range}")
        
        if table_range is None:            
            raise ValueError(f"Table '{table_name}' not found in sheet '{sheet_name}'.")
        
        # Extract data from the table range
        self.logger.debug(f"Loading data from range: {table_range}")
        from openpyxl.utils import range_boundaries
        min_col_val, min_row_val, max_col_val, max_row_val = range_boundaries(table_range)                        
            
        # Type guard: ensure all values are integers
        min_col: int = min_col_val if min_col_val is not None else 0
        min_row: int = min_row_val if min_row_val is not None else 0
        max_col: int = max_col_val if max_col_val is not None else 0
        max_row: int = max_row_val if max_row_val is not None else 0
                
        # Get headers from first row
        headers: list[str] = []
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row=min_row, column=col)
            headers.append(str(cell.value) if cell.value is not None else "")
                
        # Get data rows
        data_rows: list[list[str]] = []
        for row in range(min_row + 1, max_row + 1):
            row_data: list[str] = []
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                row_data.append(str(cell.value) if cell.value is not None else "")
            # Only add non-empty rows
            if any(cell != "" for cell in row_data):
                data_rows.append(row_data)
                
        # Create DataFrame
        if not data_rows:
            raise ValueError("No data found in the specified table range")
        df = pd.DataFrame(data_rows, columns=headers)
        self.logger.debug(f"Loaded data from table '{table_name}' with {len(df)} rows and {len(df.columns)} columns")
        return df

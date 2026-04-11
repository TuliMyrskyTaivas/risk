import pandas as pd
import numpy as np
from scipy import stats
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend

import warnings
from datetime import datetime
import os
import seaborn as sns
import sys
import argparse
import tempfile
import shutil
import msoffcrypto
import logging
from typing import Dict
warnings.filterwarnings('ignore')

from moex_data_fetcher import MOEXDataFetcher
from pdf_report import PDFReport

class PortfolioAnalyzer:
    def __init__(self, excel_path: str):
        """
        Initialize the portfolio analyzer
        
        Parameters:
        excel_path (str): Path to the password-protected Excel file
        """
        self.logger = logging.getLogger('PortfolioAnalyzer')
        self.excel_path = excel_path
        self.password = os.environ.get('PORTFOLIO_PASSWORD')
        if not self.password:
            raise ValueError("PORTFOLIO_PASSWORD environment variable not set")
        
        self.portfolio_data: pd.DataFrame 
        self.returns_data: pd.DataFrame | None = None
        self.report_data: Dict[str, pd.DataFrame] = {}
        self.moex = MOEXDataFetcher()
        
    def load_portfolio(self):
        """
        Load portfolio data from named table "Assets" in password-protected Excel file
        """
        try:            
            import io as bytes_io
            
            self.logger.debug(f"Attempting to open: {self.excel_path}")
            
            # Create a temporary file-like object
            temp = bytes_io.BytesIO()
            
            # Decrypt the file
            with open(self.excel_path, 'rb') as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=self.password)
                office_file.decrypt(temp)
            
            # Load the decrypted data using openpyxl to access named tables
            temp.seek(0)
            
            # Save decrypted content to a temporary file for openpyxl
            with open('temp_decrypted.xlsx', 'wb') as f:
                f.write(temp.getvalue())
            
            # Load workbook with openpyxl
            from openpyxl import load_workbook
            wb = load_workbook('temp_decrypted.xlsx', data_only=True)
            
            # Check if "Assets" sheet exists
            if 'Assets' not in wb.sheetnames:
                raise ValueError("Sheet 'Assets' not found in the workbook")
            
            # Get the "Assets" sheet
            sheet = wb['Assets']
            
            # Look for named table "Assets" or named range "Assets"            
            table_range = None
            
            # Method 1: Check for Excel tables (ListObjects)
            if hasattr(sheet, 'tables') and 'Assets' in sheet.tables:
                # This is an Excel table (ListObject)
                table = sheet.tables['Assets']
                table_range = table.ref
                self.logger.debug(f"Found Excel table 'Assets' at range: {table_range}")
            
            # Method 2: Check for named ranges (defined names)
            if table_range is None and hasattr(wb, 'defined_names'):
                for defined_name in wb.defined_names:
                    if defined_name.name == 'Assets':
                        # Get the range of the named range
                        destinations = defined_name.destinations
                        for title, coord in destinations:
                            if title == 'Assets':  # Range is on Assets sheet
                                table_range = coord
                                self.logger.debug(f"Found named range 'Assets' at range: {table_range}")
                                break
            
            if not table_range:
                raise ValueError("Named table or range 'Assets' not found in the workbook")

            # Extract data from the table range
            self.logger.debug(f"Loading data from range: {table_range}")
            from openpyxl.utils import range_boundaries
            min_col_val, min_row_val, max_col_val, max_row_val = range_boundaries(table_range)                        
            
            # Type guard: ensure all values are integers
            min_col: int = min_col_val if min_col_val is not None else 0
            min_row: int = min_row_val if min_row_val is not None else 0
            max_col: int = max_col_val if max_col_val is not None else 0
            max_row: int = max_row_val if max_row_val is not None else 0
                
            # Get headers from first row
            headers: list[str] = []
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=min_row, column=col)
                headers.append(str(cell.value) if cell.value is not None else "")
                
            # Get data rows
            data_rows: list[list[str]] = []
            for row in range(min_row + 1, max_row + 1):
                row_data: list[str] = []
                for col in range(min_col, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    row_data.append(str(cell.value) if cell.value is not None else "")
                # Only add non-empty rows
                if any(cell != "" for cell in row_data):
                    data_rows.append(row_data)
                
            # Create DataFrame
            if not data_rows:
                raise ValueError("No data found in the specified table range")

            self.portfolio_data = pd.DataFrame(data_rows, columns=headers)
            # Filter out rows with zero amount
            if 'Amount' in self.portfolio_data.columns:
                self.portfolio_data = self.portfolio_data[self.portfolio_data['Amount'] != '0']
            self.logger.info(f"Loaded named table/range 'Assets' with {len(self.portfolio_data)} rows")            
            self.logger.debug(f"Portfolio data:\n{self.portfolio_data.head()}")

            # Clean up temporary file
            wb.close()
            if os.path.exists('temp_decrypted.xlsx'):
                os.remove('temp_decrypted.xlsx')                    
            
        except Exception as e:
            self.logger.error(f"Error loading Excel file: {e}")
            import traceback
            self.logger.debug(traceback.print_exc())
            return False
        
        # Clean up data
        self._clean_data()
        return True        
    
    def _clean_data(self):
        """Clean and prepare the portfolio data"""
        # Strip whitespace from column names
        self.portfolio_data.columns = self.portfolio_data.columns.str.strip()
        
        # Print available columns for debugging
        self.logger.debug(f"Available columns: {list(self.portfolio_data.columns)}")
        
        # Define expected columns
        expected_columns = ['Name', 'Type', 'Code', 'Amount', 'Currency', 
                           'Book price', 'Current price', 'Book value', 
                           'Current value', 'P/L', 'Return', 'Yield']
        
        # Check which expected columns are present
        missing_columns = [col for col in expected_columns if col not in self.portfolio_data.columns]
        if missing_columns:
            self.logger.warning(f"Warning: Missing columns: {missing_columns}")
        
        # Convert numeric columns
        numeric_columns = ['Amount', 'Book price', 'Current price', 
                          'Book value', 'Current value', 'P/L', 'Return', 'Yield']
        
        for col in numeric_columns:
            if col in self.portfolio_data.columns:
                # Handle different number formats
                self.portfolio_data[col] = self.portfolio_data[col].astype(str).str.replace(',', '.')
                self.portfolio_data[col] = self.portfolio_data[col].str.replace(' ', '')
                self.portfolio_data[col] = pd.to_numeric(self.portfolio_data[col], errors='coerce')
        
        # Set currency to RUB if not specified
        if 'Currency' in self.portfolio_data.columns:
            self.portfolio_data['Currency'] = 'RUB'
        else:
            self.portfolio_data['Currency'] = 'RUB'
        
        # Update current prices from MOEX if Code column exists
        if 'Code' in self.portfolio_data.columns:
            self.logger.info("Fetching current prices from MOEX...")
            self._update_prices_from_moex()
        
        # Calculate current value based on amount and current price
        if 'Amount' in self.portfolio_data.columns and 'Current price' in self.portfolio_data.columns:
            self.portfolio_data['Current value'] = self.portfolio_data['Amount'] * self.portfolio_data['Current price']
        
        # Calculate P/L if Book value and Current value exist
        if 'Book value' in self.portfolio_data.columns and 'Current value' in self.portfolio_data.columns:
            self.portfolio_data['P/L'] = self.portfolio_data['Current value'] - self.portfolio_data['Book value']
            self.portfolio_data['P/L %'] = (self.portfolio_data['P/L'] / self.portfolio_data['Book value'] * 100).round(2)
        
        # IMPORTANT: The 'Return' column contains profit in RUB, not percentage
        # We'll keep it as is and also calculate return percentage
        if 'Return' in self.portfolio_data.columns and 'Book value' in self.portfolio_data.columns:            
            # Calculate return percentage from the RUB return
            # Avoid division by zero and handle infinite values
            mask = self.portfolio_data['Book value'] != 0
            self.portfolio_data.loc[mask, 'Return %'] = (self.portfolio_data.loc[mask, 'Return'] / 
                                                         self.portfolio_data.loc[mask, 'Book value'] * 100).round(2)
        
            # Replace any infinite values with NaN
            self.portfolio_data['Return %'] = self.portfolio_data['Return %'].replace([np.inf, -np.inf], np.nan)    
            self.logger.debug(f"Calculated return percentages from RUB values")
        
        # Calculate weights based on current value
        if 'Current value' in self.portfolio_data.columns:
            total_value = self.portfolio_data['Current value'].sum()
            self.portfolio_data['Weight'] = self.portfolio_data['Current value'] / total_value
            
            self.logger.info(f"Total Portfolio Value: {total_value:,.2f} RUB")
            
            # Store summary data
            self.report_data['total_value'] = total_value
            self.report_data['total_book_value'] = self.portfolio_data['Book value'].sum() if 'Book value' in self.portfolio_data.columns else 0
            self.report_data['total_pl'] = self.portfolio_data['P/L'].sum() if 'P/L' in self.portfolio_data.columns else 0
            
            # Total return in RUB
            if 'Return' in self.portfolio_data.columns:
                self.report_data['total_return_rub'] = self.portfolio_data['Return'].sum()
            
            # Average return percentage
            if 'Return %' in self.portfolio_data.columns:
                self.report_data['avg_return_pct'] = self.portfolio_data['Return %'].mean()
            
            self.report_data['n_assets'] = len(self.portfolio_data)
        else:
            self.logger.warning("Warning: 'Current value' column not found")
    
    def _update_prices_from_moex(self):
        """Update current prices from MOEX for all assets"""                
        updated_count = 0
        for idx, row in self.portfolio_data.iterrows():
            ticker = row['Code']
            if pd.notna(ticker) and ticker:
                price = self.moex.get_current_price(ticker)
                if price is not None:
                    self.portfolio_data.at[idx, 'Current price'] = price
                    updated_count += 1
                    self.logger.debug(f"{ticker}: {price:.2f} RUB")
                else:
                    self.logger.warning(f"Could not fetch price for {ticker}")
        
        self.logger.info(f"Updated {updated_count} prices from MOEX")
    
    def generate_returns_data(self, n_simulations: int = 10000, time_horizon : int = 1):
        """
        Generate returns data for VaR calculation.
        Prefer actual historical returns from MOEX if available.
        """
        portfolio_returns = None
        historical_returns : Dict[str, pd.Series] = {}

        if 'Code' in self.portfolio_data.columns:
            self.logger.info("Fetching historical prices for returns calculation...")
            for idx, row in self.portfolio_data.iterrows():
                ticker = row['Code']
                if pd.notna(ticker) and ticker:
                    prices = self.moex.get_historical_prices(ticker)
                    if prices is not None and len(prices) > 1:
                        returns = prices.pct_change().dropna()
                        if not returns.empty:
                            historical_returns[ticker] = returns
                            self.logger.debug(f"{ticker}: {len(returns)} historical returns loaded")                            
                        else:
                            self.logger.warning(f"{ticker}: insufficient historical price data")
                    else:
                        self.logger.warning(f"{ticker}: no historical prices available")

        if historical_returns:
            try:                
                # Handle duplicate indices by removing duplicates from each series
                clean_returns = {}
                for ticker, returns_series in historical_returns.items():
                    # Remove duplicate index values, keeping the first occurrence                    
                    clean_returns[ticker] = returns_series[~returns_series.index.duplicated(keep='first')]
                
                returns_df = pd.DataFrame(clean_returns)
                returns_df = returns_df.dropna(axis=0, how='any')
            except Exception as e:
                self.logger.warning(f"Error creating returns DataFrame: {e}. Falling back to simulation mode.")
                returns_df = pd.DataFrame()

            if not returns_df.empty:
                # Determine portfolio weights for tickers with historical data
                if 'Weight' in self.portfolio_data.columns:
                    weights = []
                    for ticker in returns_df.columns:
                        mask = self.portfolio_data['Code'] == ticker
                        weights.append(self.portfolio_data.loc[mask, 'Weight'].sum())
                    weights = np.array(weights, dtype=float)
                else:
                    weights = np.ones(len(returns_df.columns), dtype=float)

                weights = np.nan_to_num(weights, nan=0.0)
                weights_sum = weights.sum()
                if weights_sum > 0:
                    weights = weights / weights_sum
                else:
                    weights = np.ones(len(returns_df.columns), dtype=float) / len(returns_df.columns)

                portfolio_returns = returns_df.dot(weights)                                
                self.returns_data = portfolio_returns.values * np.sqrt(time_horizon)
                return self.returns_data

        # Fallback: simulate based on portfolio return percentages or Russian market assumptions
        if 'Return %' in self.portfolio_data.columns:
            asset_returns = self.portfolio_data['Return %'].dropna()
            asset_returns = asset_returns[np.isfinite(asset_returns)] / 100

            if len(asset_returns) > 5:
                mean_return = asset_returns.mean()
                std_return = asset_returns.std()
            else:
                mean_return = 0.0005
                std_return = 0.015
        else:
            mean_return = 0.0005
            std_return = 0.015

        if not np.isfinite(std_return) or std_return <= 0:
            std_return = 0.015

        n_assets = len(self.portfolio_data)
        if 'Weight' in self.portfolio_data.columns:
            weights = self.portfolio_data['Weight'].values
            weights = np.nan_to_num(weights, nan=0.0)
            weights = weights / weights.sum() if weights.sum() > 0 else np.ones(n_assets) / n_assets
        else:
            weights = np.ones(n_assets) / n_assets

        portfolio_returns = []
        for _ in range(n_simulations):
            asset_sim_returns = np.random.normal(mean_return, std_return, n_assets)
            portfolio_return = np.sum(weights * asset_sim_returns)
            portfolio_returns.append(portfolio_return)

        self.returns_data = np.array(portfolio_returns) * np.sqrt(time_horizon)
        return self.returns_data    
    
    def calculate_risk_metrics(self):
        """Calculate all risk metrics"""
        if self.returns_data is None:
            self.generate_returns_data()
        
        risk_metrics = {}
        total_value = self.report_data.get('total_value', 1)
        
        # VaR at different confidence levels
        for conf in [0.90, 0.95, 0.99]:
            var = np.percentile(self.returns_data, (1 - conf) * 100)
            risk_metrics[f'VaR_{int(conf*100)}'] = {
                'percentage': var * 100,
                'value': total_value * abs(var)
            }
        
        # Expected Shortfall
        for conf in [0.90, 0.95, 0.99]:
            var_threshold = np.percentile(self.returns_data, (1 - conf) * 100)
            tail_returns = self.returns_data[self.returns_data <= var_threshold]
            es = tail_returns.mean() if len(tail_returns) > 0 else var_threshold
            risk_metrics[f'ES_{int(conf*100)}'] = {
                'percentage': es * 100,
                'value': total_value * abs(es)
            }
        
        # Statistics
        volatility = np.std(self.returns_data) * np.sqrt(252)
        skewness = stats.skew(self.returns_data)
        kurtosis = stats.kurtosis(self.returns_data)
        
        # Sharpe ratio (using Russian risk-free rate - approximate)
        risk_free_rate = 0.15 / 252  # 15% annual key rate / 252 days
        mean_daily_return = np.mean(self.returns_data)
        sharpe_ratio = (mean_daily_return - risk_free_rate) / np.std(self.returns_data) * np.sqrt(252)
        
        # Maximum drawdown
        cumulative_returns = np.cumprod(1 + self.returns_data)
        running_max = np.maximum.accumulate(cumulative_returns)
        drawdown = (cumulative_returns - running_max) / running_max
        max_drawdown = drawdown.min() * 100
        
        risk_metrics['statistics'] = {
            'volatility': volatility * 100,
            'skewness': skewness,
            'kurtosis': kurtosis,
            'sharpe_ratio': sharpe_ratio,
            'max_drawdown': max_drawdown
        }
        
        self.report_data['risk_metrics'] = risk_metrics
        return risk_metrics
    
    def create_visualizations(self):
        """Create all visualizations for the report"""
        vis_data = {}
    
        # Set style
        plt.style.use('seaborn-v0_8-darkgrid')
        sns.set_palette("husl")
    
        # 1. Asset Allocation Pie Chart
        if 'Type' in self.portfolio_data.columns and 'Current value' in self.portfolio_data.columns:
            fig1, ax1 = plt.subplots(figsize=(10, 8))
            type_allocation = self.portfolio_data.groupby('Type')['Current value'].sum()
            colors = plt.cm.Set3(np.linspace(0, 1, len(type_allocation)))
            wedges, texts, autotexts = ax1.pie(type_allocation.values, 
                                                labels=type_allocation.index,
                                                autopct='%1.1f%%',
                                                colors=colors,
                                                startangle=90)
            ax1.set_title('Asset Allocation by Type', fontsize=16, fontweight='bold')
            plt.setp(autotexts, size=10, weight="bold")
            vis_data['allocation_pie'] = fig1
    
        # 2. Top 10 Assets by Value
        if 'Name' in self.portfolio_data.columns and 'Current value' in self.portfolio_data.columns:
            fig2, ax2 = plt.subplots(figsize=(12, 6))
            top_10 = self.portfolio_data.nlargest(10, 'Current value')[['Name', 'Current value', 'Type']]
            bars = ax2.barh(range(len(top_10)), top_10['Current value'].values)
            ax2.set_yticks(range(len(top_10)))
            ax2.set_yticklabels(top_10['Name'].values)
            ax2.set_xlabel('Current Value (RUB)', fontsize=12)
            ax2.set_title('Top 10 Assets by Value', fontsize=16, fontweight='bold')
        
            # Add value labels
            for i, (bar, val) in enumerate(zip(bars, top_10['Current value'].values)):
                ax2.text(val, i, f' {val:,.0f} RUB', va='center', fontweight='bold')
        
            plt.tight_layout()
            vis_data['top_assets'] = fig2
    
        # 3. Returns Distribution (using Return %) - FIXED VERSION
        if 'Return %' in self.portfolio_data.columns:
            # Clean the returns data - remove inf and NaN values
            returns = self.portfolio_data['Return %'].dropna()
            returns = returns[np.isfinite(returns)]  # Remove any infinite values
        
            if len(returns) > 0:  # Only create histogram if we have valid data
                fig3, (ax3, ax4) = plt.subplots(1, 2, figsize=(14, 6))
            
                # Histogram of returns
                ax3.hist(returns, bins=20, edgecolor='black', alpha=0.7)
                ax3.axvline(returns.mean(), color='red', linestyle='dashed', linewidth=2, 
                           label=f'Average: {returns.mean():.2f}%')
                ax3.axvline(0, color='black', linewidth=1)
                ax3.set_xlabel('Returns (%)')
                ax3.set_ylabel('Frequency')
                ax3.set_title('Distribution of Asset Returns (%)', fontsize=16, fontweight='bold')
                ax3.legend()
            
                # Box plot by type
                if 'Type' in self.portfolio_data.columns and len(self.portfolio_data['Type'].unique()) > 1:
                    type_returns = []
                    type_labels = []
                    for t in self.portfolio_data['Type'].unique():
                        # Clean data for each type
                        type_data = self.portfolio_data[self.portfolio_data['Type'] == t]['Return %'].dropna()
                        type_data = type_data[np.isfinite(type_data)]
                        if len(type_data) > 0:
                            type_returns.append(type_data)
                            type_labels.append(t)
                
                    if type_returns:  # Only create boxplot if we have valid data
                        ax4.boxplot(type_returns, labels=type_labels)
                        ax4.set_ylabel('Returns (%)')
                        ax4.set_title('Returns by Asset Type', fontsize=16, fontweight='bold')
                        ax4.tick_params(axis='x', rotation=45)
            
                plt.tight_layout()
                vis_data['returns_dist'] = fig3
            else:
                self.logger.warning("Warning: No valid return percentage data for histogram")
    
        # 4. Profit/Loss Analysis (using Return in RUB) - FIXED VERSION
        if 'Return' in self.portfolio_data.columns:
            # Clean the return data
            returns_rub = self.portfolio_data['Return'].dropna()
            returns_rub = returns_rub[np.isfinite(returns_rub)]
        
            if len(returns_rub) > 0:
                fig4, ax5 = plt.subplots(figsize=(12, 6))
            
                # Get top profitable and top unprofitable
                profitable = self.portfolio_data[self.portfolio_data['Return'] > 0].nlargest(15, 'Return')[['Name', 'Return']]
                unprofitable = self.portfolio_data[self.portfolio_data['Return'] < 0].nsmallest(15, 'Return')[['Name', 'Return']]
            
                if len(profitable) > 0 or len(unprofitable) > 0:
                    pl_sorted = pd.concat([profitable, unprofitable])
                
                    if len(pl_sorted) > 0:
                        colors = ['green' if x > 0 else 'red' for x in pl_sorted['Return']]
                        bars = ax5.barh(range(len(pl_sorted)), pl_sorted['Return'].values, color=colors)
                        ax5.set_yticks(range(len(pl_sorted)))
                        ax5.set_yticklabels(pl_sorted['Name'].values)
                        ax5.set_xlabel('Profit/Loss (RUB)', fontsize=12)
                        ax5.set_title('Top 15 Profitable and Unprofitable Assets', fontsize=16, fontweight='bold')
                    
                        # Add value labels
                        for i, (bar, val) in enumerate(zip(bars, pl_sorted['Return'].values)):
                            ax5.text(val, i, f' {val:,.0f} RUB', va='center', fontweight='bold')
                    
                        plt.tight_layout()
                        vis_data['profit_loss'] = fig4
    
        self.report_data['visualizations'] = vis_data
        return vis_data

def generate_pdf_report(analyzer: PortfolioAnalyzer, filename: str = "portfolio_analysis_report.pdf"):
    """Generate comprehensive PDF report"""
    
    logger = logging.getLogger('PortfolioAnalyzer')

    # Create a temporary directory for images
    temp_dir = tempfile.mkdtemp()
    image_files = []    
    
    try:
        # Create PDF
        pdf = PDFReport()
        pdf.add_page()
        
        # Executive Summary
        pdf.chapter_title("Executive Summary")
        
        total_return_rub = analyzer.report_data.get('total_return_rub', 0)
        total_value = analyzer.report_data.get('total_value', 0)
        return_pct = (total_return_rub / (analyzer.report_data.get('total_book_value', 1)) * 100) if analyzer.report_data.get('total_book_value', 0) > 0 else 0
        
        summary_text = f"""
        This report provides a comprehensive analysis of the risks associated with your investment portfolio
        as of {datetime.now().strftime('%d %B %Y')}.
        
        Executive Summary:
        - Total Portfolio Value: {analyzer.report_data.get('total_value', 0):,.2f} RUB
        - Total Book Value: {analyzer.report_data.get('total_book_value', 0):,.2f} RUB
        - Total P&L: {analyzer.report_data.get('total_pl', 0):,.2f} RUB
        - Total Return (RUB): {total_return_rub:,.2f} RUB
        - Total Return (%): {return_pct:.2f}%
        - Number of Assets: {analyzer.report_data.get('n_assets', 0)}
        """
        pdf.chapter_body(summary_text)
        
        # Risk Metrics
        if 'risk_metrics' in analyzer.report_data:
            pdf.chapter_title("Risk Metrics Analysis")
            
            risk_metrics = analyzer.report_data['risk_metrics']
            risk_text = f"""
            Value at Risk (VaR) - Maximum Expected Loss over 1 day at different confidence levels:
            - 90% VaR: {risk_metrics['VaR_90']['percentage']:.2f}% ({risk_metrics['VaR_90']['value']:,.2f} RUB)
            - 95% VaR: {risk_metrics['VaR_95']['percentage']:.2f}% ({risk_metrics['VaR_95']['value']:,.2f} RUB)
            - 99% VaR: {risk_metrics['VaR_99']['percentage']:.2f}% ({risk_metrics['VaR_99']['value']:,.2f} RUB)
            
            Expected Shortfall (Expected Shortfall - ES) - Average Loss in the Worst Cases:
            - 90% ES: {risk_metrics['ES_90']['percentage']:.2f}% ({risk_metrics['ES_90']['value']:,.2f} RUB)
            - 95% ES: {risk_metrics['ES_95']['percentage']:.2f}% ({risk_metrics['ES_95']['value']:,.2f} RUB)
            - 99% ES: {risk_metrics['ES_99']['percentage']:.2f}% ({risk_metrics['ES_99']['value']:,.2f} RUB)
            """
            pdf.chapter_body(risk_text)
            
            # Statistical Measures
            if 'statistics' in risk_metrics:
                stats = risk_metrics['statistics']
                stats_text = f"""
                Statistical Measures:
                - Annual Volatility: {stats['volatility']:.2f}%
                - Skewness: {stats['skewness']:.3f}
                - Kurtosis: {stats['kurtosis']:.3f}
                - Sharpe Ratio: {stats['sharpe_ratio']:.2f}
                - Maximum Drawdown: {stats['max_drawdown']:.2f}%
                """
                pdf.chapter_body(stats_text)
        
        # Add visualizations
        if 'visualizations' in analyzer.report_data:
            vis_data = analyzer.report_data['visualizations']
            
            # Asset Allocation
            if 'allocation_pie' in vis_data:
                pdf.chapter_title("Asset Allocation Analysis")
                pdf.cell(0, 10, "Portfolio Distribution by Asset Types:", 0, 1)
                
                fig = vis_data['allocation_pie']
                img_path = os.path.join(temp_dir, 'allocation_pie.png')
                fig.savefig(img_path, format='png', dpi=150, bbox_inches='tight')
                image_files.append(img_path)
                
                pdf.image(img_path, x=30, y=None, w=150)
                pdf.ln(80)
                plt.close(fig)
            
            # Top 10 assets
            if 'top_assets' in vis_data:
                pdf.add_page()
                pdf.chapter_title("Top 10 Assets by Value")
                
                fig = vis_data['top_assets']
                img_path = os.path.join(temp_dir, 'top_assets.png')
                fig.savefig(img_path, format='png', dpi=150, bbox_inches='tight')
                image_files.append(img_path)
                
                pdf.image(img_path, x=20, y=None, w=170)
                pdf.ln(70)
                plt.close(fig)
            
            # Returns Distribution
            if 'returns_dist' in vis_data:
                pdf.add_page()
                pdf.chapter_title("Returns Distribution Analysis")
                
                fig = vis_data['returns_dist']
                img_path = os.path.join(temp_dir, 'returns_dist.png')
                fig.savefig(img_path, format='png', dpi=150, bbox_inches='tight')
                image_files.append(img_path)
                
                pdf.image(img_path, x=20, y=None, w=170)
                pdf.ln(70)
                plt.close(fig)
            
            # Profit/Loss Analysis
            if 'profit_loss' in vis_data:
                pdf.add_page()
                pdf.chapter_title("Profit/Loss Analysis")
                
                fig = vis_data['profit_loss']
                img_path = os.path.join(temp_dir, 'profit_loss.png')
                fig.savefig(img_path, format='png', dpi=150, bbox_inches='tight')
                image_files.append(img_path)
                
                pdf.image(img_path, x=20, y=None, w=170)
                pdf.ln(70)
                plt.close(fig)
        
        # Asset Type Analysis Table
        if analyzer.portfolio_data is not None and 'Type' in analyzer.portfolio_data.columns:
            pdf.add_page()
            pdf.chapter_title("Asset Type Analysis")
            
            type_analysis = analyzer.portfolio_data.groupby('Type').agg({
                'Current value': 'sum' if 'Current value' in analyzer.portfolio_data.columns else lambda x: 0,
                'Return': 'sum' if 'Return' in analyzer.portfolio_data.columns else lambda x: 0,  # Sum of RUB returns
                'Return %': 'mean' if 'Return %' in analyzer.portfolio_data.columns else lambda x: 0,
                'P/L': 'sum' if 'P/L' in analyzer.portfolio_data.columns else lambda x: 0
            }).round(2)
            
            # Format the data for display
            display_data = type_analysis.copy()
            if 'Current value' in display_data.columns:
                display_data['Current value'] = display_data['Current value'].apply(lambda x: f"{x:,.0f} RUB")
            if 'Return' in display_data.columns:
                display_data['Return'] = display_data['Return'].apply(lambda x: f"{x:,.0f} RUB")
            if 'Return %' in display_data.columns:
                display_data['Return %'] = display_data['Return %'].apply(lambda x: f"{x:.2f}%")
            if 'P/L' in display_data.columns:
                display_data['P/L'] = display_data['P/L'].apply(lambda x: f"{x:,.0f} RUB")
            
            display_data = display_data.reset_index()
            
            # Calculate column widths
            col_widths = [35, 40, 35, 30, 30]  # Adjusted for Russian text
            pdf.add_table("Asset Type Metrics", display_data, col_widths)
        
        # Detailed Asset Table
        if analyzer.portfolio_data is not None:
            pdf.add_page()
            pdf.chapter_title("Detailed Asset List")
            
            # Select columns to display
            display_cols = []
            for col in ['Name', 'Type', 'Code', 'Amount', 'Current price', 
                       'Current value', 'Return', 'Return %', 'P/L', 'Weight']:
                if col in analyzer.portfolio_data.columns:
                    display_cols.append(col)
            
            if display_cols:
                asset_table = analyzer.portfolio_data[display_cols].copy()
                asset_table = asset_table.head(50)  # Limit to 50 rows for PDF
                
                # Format columns
                formatted_table = asset_table.copy()
                if 'Current price' in formatted_table.columns:
                    formatted_table['Current price'] = formatted_table['Current price'].apply(
                        lambda x: f"{x:,.2f} RUB" if pd.notna(x) else "N/A")
                if 'Current value' in formatted_table.columns:
                    formatted_table['Current value'] = formatted_table['Current value'].apply(
                        lambda x: f"{x:,.0f} RUB" if pd.notna(x) else "N/A")
                if 'Return' in formatted_table.columns:
                    formatted_table['Return'] = formatted_table['Return'].apply(
                        lambda x: f"{x:,.0f} RUB" if pd.notna(x) else "N/A")
                if 'Return %' in formatted_table.columns:
                    formatted_table['Return %'] = formatted_table['Return %'].apply(
                        lambda x: f"{x:.2f}%" if pd.notna(x) else "N/A")
                if 'P/L' in formatted_table.columns:
                    formatted_table['P/L'] = formatted_table['P/L'].apply(
                        lambda x: f"{x:,.0f} RUB" if pd.notna(x) else "N/A")
                if 'Weight' in formatted_table.columns:
                    formatted_table['Weight'] = formatted_table['Weight'].apply(
                        lambda x: f"{x*100:.2f}%" if pd.notna(x) else "N/A")
                
                # Calculate column widths
                col_widths = [pdf.w / len(display_cols) - 3] * len(display_cols)
                pdf.add_table("Detailed Asset List (Top 50)", formatted_table, col_widths)
        
        # Save PDF
        pdf.output(filename)
        logger.info(f"PDF report generated: {filename}")
        
    finally:
        # Clean up temporary image files
        for img_file in image_files:
            try:
                if os.path.exists(img_file):
                    os.remove(img_file)
            except:
                pass
        
        # Remove temporary directory
        try:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
        except:
            pass        

def main():
    """
    Main function to run the portfolio analysis
    """
    parser = argparse.ArgumentParser(description='Analyze portfolio risk from a password-protected Excel file and generate a PDF report')
    parser.add_argument('excel_file', help='Path to the password-protected Excel file containing the portfolio data')
    parser.add_argument('--verbose', '-v', action='store_true', help='Enable verbose output for debugging')
    parser.add_argument('--output', '-o', default='portfolio_risk_report.pdf', 
                       help='Name of the output PDF file (default: portfolio_risk_report.pdf)')
    
    args = parser.parse_args()
    
    # Setup logging
    logger = logging.getLogger('PortfolioAnalyzer')
    handler = logging.StreamHandler()        

    if args.verbose:
        logger.setLevel(logging.DEBUG)        
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
    else:
        logger.setLevel(logging.INFO)
        formatter = logging.Formatter('%(message)s')
        handler.setFormatter(formatter)

    logger.addHandler(handler)

    logger.info("Starting portfolio analysis...")
    logger.debug(f"Input Excel file: {args.excel_file}, output PDF: {args.output}")    
            
    try:
        # Create analyzer instance
        analyzer = PortfolioAnalyzer(args.excel_file)
        
        # Load portfolio data
        if analyzer.load_portfolio():
            logger.info("Calculating risk metrics...")
            analyzer.calculate_risk_metrics()
            
            logger.info("Creating visualizations...")
            analyzer.create_visualizations()
            
            logger.info("Generating PDF report...")
            generate_pdf_report(analyzer, args.output)
            
            logger.info(f"Analysis completed, generated report: {args.output}")            
    
    except Exception as e:
        logger.error(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()